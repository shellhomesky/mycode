import openpyxl
import pandas as pd  # 存取csv
import requests  # 发送请求
from bs4 import BeautifulSoup  # 解析网页
from lxml import etree  # 导入xpath包

# 加载工作薄
wb = openpyxl.Workbook()
# 如果不存在即创建
wb.save(r'DoubleBall.xlsx')

# 打开excel工作簿
wb = openpyxl.load_workbook(r'../data/DoubleBall.xlsx')
# 准备写工作簿中第一个sheet页
sheet1 = wb.worksheets[0]

# 为数据页循环标识
indexnum = 1
print('爬取开始了')
# http://kaijiang.zhcw.com/zhcw/html/ssq/list_2.html
# 初始化URL第一页地址
url = 'http://kaijiang.zhcw.com/zhcw/html/ssq/list.html'
# 请求头部，伪造浏览器，防止爬虫被反
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'
}
# 定义excel中的行数
indexcount = 0
# 初步了解数据总页数为151，可以放更大的值
while indexnum < 152:
    # 如果不是数据页第一页则拼接数据页地址
    if indexnum > 1:
        url = 'http://kaijiang.zhcw.com/zhcw/html/ssq/list_' + str(indexnum) + '.html'

    # 利用请求地址和请求头部构造请求对象
    res = requests.get(url, headers=headers)
    # BeautifulSoup解析包
    soup = BeautifulSoup(res.text, 'html.parser')
    # xpath解析包 构造 _Element 对象
    res_elements = etree.HTML(res.text)

    # 使用 xpath 匹配数据，找到页面的table列表标签 ，
    table = res_elements.xpath('//table')
    # 函数可以将转换为Element对象再转换回html字符串
    table = etree.tostring(table[0], encoding='utf-8').decode()
    # read_html()函数是将HTML的表格转换为DataFrame
    df = pd.read_html(table, encoding='utf-8', header=0)[0]
    # 将dataframe进行转置，进而用to_dict()方法转为字典类型
    results = list(df.T.to_dict().values())

    # 根据每页数据条数，控制循环次数
    count = 1
    # 循环取每页的数据
    for index in results:
        # 只有第一页和非第一页且不是第一条数据需要取出来存放到excel(因为每页都有表头)
        if indexnum == 1 or (indexnum != 1 and count != 1):
            # 如果一页的数据循环到最后一条时退出，跳转到下一页
            if count < len(results):
                # 定义并赋值数据所需要存储的行号
                indexcount = (indexnum - 1) * 20 + count
                sheet1.cell(indexcount, 1, index.get('开奖日期'))
                sheet1.cell(indexcount, 2, index.get('期号'))
                # 由于所有号码都是存在一个值当中，所以需要通过按空格分隔为list
                listball = index.get('中奖号码').split()
                # 判断数据是否第一行是否为表头，如果大于1就不是，则动态赋值，如果是则直接写文字内容
                if len(listball) > 1:
                    sheet1.cell(indexcount, 3, listball[0])
                    sheet1.cell(indexcount, 4, listball[1])
                    sheet1.cell(indexcount, 5, listball[2])
                    sheet1.cell(indexcount, 6, listball[3])
                    sheet1.cell(indexcount, 7, listball[4])
                    sheet1.cell(indexcount, 8, listball[5])
                    sheet1.cell(indexcount, 9, listball[6])
                else:
                    sheet1.cell(indexcount, 3, '红球1')
                    sheet1.cell(indexcount, 4, '红球2')
                    sheet1.cell(indexcount, 5, '红球3')
                    sheet1.cell(indexcount, 6, '红球4')
                    sheet1.cell(indexcount, 7, '红球5')
                    sheet1.cell(indexcount, 8, '红球6')
                    sheet1.cell(indexcount, 9, '蓝球')

                sheet1.cell(indexcount, 10, index.get('中奖注数'))
                sheet1.cell(indexcount, 11, index.get('中奖注数.1'))

        count += 1

    indexnum += 1

# 保存关闭excel
wb.save(r'DoubleBall.xlsx')
print('爬取结束了')
